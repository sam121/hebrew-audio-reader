from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from common import (
    BALANCE_COLUMNS,
    CONTROL_TOTAL_COLUMNS,
    HOLDING_COLUMNS,
    ISSUE_COLUMNS,
    PROCESSED_DIR,
    converted_with_fx,
    ensure_dirs,
    make_issue,
    owner_for_path,
    source_folders_for,
    stable_id,
    write_csv,
)


PARSER_NAME = "ingest_vanguard_inferred_v1"


def vanguard_files() -> list[Path]:
    candidates: list[Path] = []
    for root in source_folders_for("vanguard"):
        candidates.extend(list(root.glob("*.xlsx")) + list(root.glob("*.Xlsx")) + list(root.glob("*.XLSX")))
    return sorted(path for path in candidates if not path.name.startswith("~$") and path.stat().st_size > 1024)


def account_context(title: str, first_cell: Any) -> tuple[str, str, str]:
    account_name = str(first_cell or title).strip()
    if "(" in title and ")" in title:
        account_id = title.split("(", 1)[1].split(")", 1)[0]
    elif "Pension" in account_name:
        account_id = "VANGUARD_PERSONAL_PENSION"
    else:
        account_id = title
    account_type = "pension" if "Pension" in account_name else ("isa" if "ISA" in account_name else "investment")
    return account_id, account_name, account_type


def run() -> dict[str, Any]:
    ensure_dirs()
    try:
        from openpyxl import load_workbook
    except ImportError:
        write_csv(PROCESSED_DIR / "vanguard_inferred_balances.csv", [], BALANCE_COLUMNS)
        write_csv(PROCESSED_DIR / "vanguard_inferred_holdings.csv", [], HOLDING_COLUMNS)
        write_csv(PROCESSED_DIR / "vanguard_inferred_parse_issues.csv", [], ISSUE_COLUMNS)
        write_csv(PROCESSED_DIR / "vanguard_inferred_control_totals.csv", [], CONTROL_TOTAL_COLUMNS)
        return {"files": 0, "balances": 0, "holdings": 0, "issues": 1}

    balances: list[dict[str, Any]] = []
    holdings: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    controls: list[dict[str, Any]] = []

    files = vanguard_files()
    for path in files:
      owner = owner_for_path(path)
      wb = load_workbook(path, read_only=True, data_only=True)
      for ws in wb.worksheets:
        if ws.title == "Summary":
            continue
        account_id, account_name, account_type = account_context(ws.title, ws.cell(row=1, column=1).value)
        current = None
        events = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = list(row)
            if vals[0] == "Date" and len(vals) > 1 and vals[1] == "InvestmentName":
                current = "investment"
                continue
            if current == "investment" and isinstance(vals[0], datetime):
                events.append(
                    {
                        "source_row": idx,
                        "date": vals[0].date(),
                        "fund": str(vals[1] or "").strip(),
                        "quantity": Decimal(str(vals[3] or 0)),
                        "price": Decimal(str(vals[4] or 0)),
                        "cost": Decimal(str(vals[5] or 0)),
                    }
                )
        events.sort(key=lambda item: item["date"])
        if events and all(event["quantity"] <= 0 for event in events):
            issues.append(
                make_issue(
                    "valuation",
                    "warning",
                    owner,
                    "vanguard",
                    account_id,
                    "Vanguard investment rows only show sales/fee funding, so the workbook does not contain enough history to infer a current portfolio value.",
                    "Add a Vanguard valuation/holdings export or a manual balance point for this account.",
                    value_date=max(event["date"] for event in events),
                    source_file=str(path.resolve()),
                    source_page=ws.title,
                    key=f"vanguard-inferred-negative-only-{owner}-{path.name}-{ws.title}",
                )
            )
            controls.append(
                {
                    "control_id": stable_id("ctl", PARSER_NAME, str(path.resolve()), ws.title),
                    "parser_name": PARSER_NAME,
                    "source_file": str(path.resolve()),
                    "owner": owner,
                    "institution": "vanguard",
                    "account_id": account_id,
                    "file_count": 1,
                    "row_count": len(events),
                    "date_min": min((event["date"] for event in events), default=None),
                    "date_max": max((event["date"] for event in events), default=None),
                    "sum_credits": None,
                    "sum_debits": None,
                    "opening_balance": None,
                    "closing_balance": None,
                    "warning_count": 1,
                    "failed_row_count": 0,
                }
            )
            continue
        units: dict[str, Decimal] = defaultdict(Decimal)
        prices: dict[str, Decimal] = {}
        price_dates: dict[str, Any] = {}
        for event in events:
            if not event["fund"]:
                continue
            units[event["fund"]] += event["quantity"]
            if event["price"] > 0:
                prices[event["fund"]] = event["price"]
                price_dates[event["fund"]] = event["date"]
            total_value = sum(qty * prices.get(fund, Decimal("0")) for fund, qty in units.items())
            value_sgd, fx = converted_with_fx(total_value, "GBP", event["date"])
            balances.append(
                {
                    "balance_id": stable_id("bal", PARSER_NAME, owner, account_id, event["date"], total_value, event["source_row"]),
                    "owner": owner,
                    "institution": "vanguard",
                    "account_id": account_id,
                    "account_name": account_name,
                    "account_type": account_type,
                    "date": event["date"],
                    "balance": total_value,
                    "currency": "GBP",
                    "balance_sgd": value_sgd,
                    **fx,
                    "balance_type": "inferred_investment_value",
                    "confidence_status": "inferred",
                    "source_file": str(path.resolve()),
                    "source_page": ws.title,
                    "source_row": event["source_row"],
                    "parser_name": PARSER_NAME,
                    "parse_confidence": 0.72,
                }
            )

        latest_date = max((event["date"] for event in events), default=None)
        for fund, quantity in sorted(units.items()):
            price = prices.get(fund, Decimal("0"))
            market_value = quantity * price
            market_value_sgd, fx = converted_with_fx(market_value, "GBP", latest_date)
            holdings.append(
                {
                    "holding_id": stable_id("hold", PARSER_NAME, owner, account_id, fund, latest_date),
                    "owner": owner,
                    "institution": "vanguard",
                    "account_id": account_id,
                    "date": latest_date,
                    "symbol": "",
                    "name": fund,
                    "asset_class": "fund",
                    "quantity": quantity,
                    "price": price,
                    "market_value": market_value,
                    "currency": "GBP",
                    "market_value_sgd": market_value_sgd,
                    **fx,
                    "confidence_status": "inferred",
                    "source_file": str(path.resolve()),
                    "source_row": ws.title,
                    "parser_name": PARSER_NAME,
                }
            )

        controls.append(
            {
                "control_id": stable_id("ctl", PARSER_NAME, str(path.resolve()), ws.title),
                "parser_name": PARSER_NAME,
                "source_file": str(path.resolve()),
                "owner": owner,
                "institution": "vanguard",
                "account_id": account_id,
                "file_count": 1,
                "row_count": len(events),
                "date_min": min((event["date"] for event in events), default=None),
                "date_max": max((event["date"] for event in events), default=None),
                "sum_credits": None,
                "sum_debits": None,
                "opening_balance": None,
                "closing_balance": balances[-1]["balance"] if balances else None,
                "warning_count": 0,
                "failed_row_count": 0,
            }
        )

    write_csv(PROCESSED_DIR / "vanguard_inferred_balances.csv", balances, BALANCE_COLUMNS)
    write_csv(PROCESSED_DIR / "vanguard_inferred_holdings.csv", holdings, HOLDING_COLUMNS)
    write_csv(PROCESSED_DIR / "vanguard_inferred_parse_issues.csv", issues, ISSUE_COLUMNS)
    write_csv(PROCESSED_DIR / "vanguard_inferred_control_totals.csv", controls, CONTROL_TOTAL_COLUMNS)
    return {"files": len(files), "balances": len(balances), "holdings": len(holdings), "issues": len(issues)}


if __name__ == "__main__":
    print(run())
