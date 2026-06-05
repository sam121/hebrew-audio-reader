from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from common import REPORTS_DIR, REPORT_END_DATE, REPORT_START_DATE, SOURCE_ROOT, converted_with_fx, html_escape, write_html_report


VANGUARD_FILE = SOURCE_ROOT / "Sam" / "sam_vanguard" / "LoadDocstore.Xlsx"


def money(value: Decimal | float) -> str:
    return f"{float(value):,.2f}"


def account_id_from_title(title: str, account_name: str) -> str:
    if "(" in title and ")" in title:
        return title.split("(", 1)[1].split(")", 1)[0]
    if "Pension" in account_name:
        return "VANGUARD_PERSONAL_PENSION"
    return title


def month_key(value: str) -> str:
    return value[:7]


def monthly_last(points: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_month: dict[str, dict[str, Any]] = {}
    for point in points:
        if point["date"] < REPORT_START_DATE.isoformat() or point["date"] > REPORT_END_DATE.isoformat():
            continue
        by_month[month_key(point["date"])] = {**point, "date": month_key(point["date"])}
    return [by_month[key] for key in sorted(by_month)]


def extract_series() -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    from openpyxl import load_workbook

    wb = load_workbook(VANGUARD_FILE, read_only=True, data_only=True)
    account_series: dict[str, list[dict[str, Any]]] = {}
    holdings_rows: list[dict[str, Any]] = []
    all_dates: set[str] = set()

    for ws in wb.worksheets:
        if ws.title == "Summary":
            continue
        account_name = str(ws.cell(row=1, column=1).value or ws.title).strip()
        account_id = account_id_from_title(ws.title, account_name)
        label = f"{account_name} ({account_id})"

        current = None
        events = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = list(row)
            if vals[0] == "Date" and len(vals) > 1 and vals[1] == "InvestmentName":
                current = "investment"
                continue
            if current == "investment" and isinstance(vals[0], datetime):
                events.append(
                    {
                        "date": vals[0].date(),
                        "fund": str(vals[1] or "").strip(),
                        "details": str(vals[2] or "").strip(),
                        "quantity": Decimal(str(vals[3] or 0)),
                        "price": Decimal(str(vals[4] or 0)),
                        "cost": Decimal(str(vals[5] or 0)),
                    }
                )

        events.sort(key=lambda item: item["date"])
        units: dict[str, Decimal] = defaultdict(Decimal)
        prices: dict[str, Decimal] = {}
        points: list[dict[str, Any]] = []
        for event in events:
            if event["fund"]:
                units[event["fund"]] += event["quantity"]
                if event["price"] > 0:
                    prices[event["fund"]] = event["price"]
            gbp_value = sum(quantity * prices.get(fund, Decimal("0")) for fund, quantity in units.items())
            sgd_value, fx = converted_with_fx(gbp_value, "GBP", event["date"])
            point = {
                "date": event["date"].isoformat(),
                "gbp": float(gbp_value),
                "sgd": float(sgd_value or 0),
            }
            points.append(point)
            if REPORT_START_DATE <= event["date"] <= REPORT_END_DATE:
                all_dates.add(point["date"])

        account_series[label] = monthly_last(points)
        for fund, quantity in sorted(units.items()):
            latest_price = prices.get(fund, Decimal("0"))
            value = quantity * latest_price
            sgd_value, fx = converted_with_fx(value, "GBP", events[-1]["date"] if events else None)
            holdings_rows.append(
                {
                    "account": label,
                    "fund": fund,
                    "units": float(quantity),
                    "latest_price_gbp": float(latest_price),
                    "value_gbp": float(value),
                    "value_sgd": float(sgd_value or 0),
                    "valuation_date": events[-1]["date"].isoformat() if events else "",
                }
            )

    latest_by_account: dict[str, dict[str, Any] | None] = {}
    for label, points in account_series.items():
        latest_by_account[label] = None
        idx = 0
        for d in sorted(all_dates):
            while idx < len(points) and points[idx]["date"] <= d:
                latest_by_account[label] = points[idx]
                idx += 1

    total_points = []
    all_months = sorted({month_key(d) for d in all_dates})
    for d in all_months:
        total_gbp = Decimal("0")
        total_sgd = Decimal("0")
        for points in account_series.values():
            prior = None
            for point in points:
                if point["date"] <= d:
                    prior = point
                else:
                    break
            if prior:
                total_gbp += Decimal(str(prior["gbp"]))
                total_sgd += Decimal(str(prior["sgd"]))
        total_points.append({"date": d, "gbp": float(total_gbp), "sgd": float(total_sgd)})

    return total_points, account_series, holdings_rows


def run() -> dict[str, Any]:
    total, series, holdings = extract_series()
    payload = json.dumps({"total": total, "series": series})
    latest = total[-1] if total else {"date": "", "gbp": 0, "sgd": 0}
    body = f"""
<div class="metric-row">
  <div class="metric"><strong>{html_escape(latest['date'])}</strong><span>Latest inferred valuation date</span></div>
  <div class="metric"><strong>GBP {money(latest['gbp'])}</strong><span>Latest inferred Vanguard value</span></div>
  <div class="metric"><strong>SGD {money(latest['sgd'])}</strong><span>Converted with ECB FX table</span></div>
  <div class="metric"><strong>{len(holdings)}</strong><span>Current holding lines</span></div>
</div>
<p class="warning">This chart stops at the last completed month-end ({REPORT_END_DATE.isoformat()}) and is inferred from Vanguard investment transaction rows, aggregated to month-end snapshots using the last inferred value available in each month. It is for sanity checking, not a confirmed Vanguard valuation export.</p>
<svg id="chart" viewBox="0 0 1100 540" role="img" aria-label="Inferred Vanguard value over time" style="width:100%;height:auto;background:#fff;border:1px solid var(--line);border-radius:6px"></svg>
<h2>Latest Holdings</h2>
<table><thead><tr><th>Account</th><th>Fund</th><th>Units</th><th>Latest Price GBP</th><th>Value GBP</th><th>Value SGD</th></tr></thead><tbody>
{''.join(f"<tr><td>{html_escape(h['account'])}</td><td>{html_escape(h['fund'])}</td><td>{h['units']:,.4f}</td><td>{h['latest_price_gbp']:,.2f}</td><td>{h['value_gbp']:,.2f}</td><td>{h['value_sgd']:,.2f}</td></tr>" for h in holdings)}
</tbody></table>
<h2>Series</h2>
<div id="legend" class="metric-row"></div>
<script>
const data = {payload};
const svg = document.getElementById('chart');
const legend = document.getElementById('legend');
const W = 1100, H = 540, L = 76, R = 32, T = 34, B = 66;
const colors = ['#116466', '#c44536', '#315f9d'];
const parseDate = d => new Date(d.length === 7 ? d + '-01T00:00:00' : d + 'T00:00:00');
const all = data.total.map(p => ({{...p, t: parseDate(p.date).getTime(), value:p.sgd}}));
const minT = Math.min(...all.map(p => p.t));
const maxT = Math.max(...all.map(p => p.t));
const vals = [];
for (const p of all) vals.push(p.value);
for (const pts of Object.values(data.series)) for (const p of pts) vals.push(p.sgd);
const minY = 0;
const maxY = Math.max(...vals);
const padY = maxY * 0.08 || 1;
const y1 = maxY + padY;
const x = t => L + (t - minT) / (maxT - minT || 1) * (W - L - R);
const y = v => T + (y1 - v) / (y1 || 1) * (H - T - B);
const el = (name, attrs, text) => {{
  const node = document.createElementNS('http://www.w3.org/2000/svg', name);
  for (const [k,v] of Object.entries(attrs || {{}})) node.setAttribute(k, v);
  if (text !== undefined) node.textContent = text;
  svg.appendChild(node);
  return node;
}};
const fmt = v => 'S$' + Math.round(v).toLocaleString();
for (let i=0;i<=5;i++) {{
  const val = y1*i/5;
  const yy = y(val);
  el('line', {{x1:L, y1:yy, x2:W-R, y2:yy, stroke:'#d9e0e7', 'stroke-width':1}});
  el('text', {{x:L-10, y:yy+4, 'text-anchor':'end', 'font-size':12, fill:'#5f6b7a'}}, fmt(val));
}}
const years = [...new Set(all.map(p => new Date(p.t).getFullYear()))];
for (const yr of years) {{
  const tt = new Date(yr + '-01-01T00:00:00').getTime();
  if (tt < minT || tt > maxT) continue;
  const xx = x(tt);
  el('line', {{x1:xx, y1:T, x2:xx, y2:H-B, stroke:'#edf1f4', 'stroke-width':1}});
  el('text', {{x:xx, y:H-34, 'text-anchor':'middle', 'font-size':12, fill:'#5f6b7a'}}, yr);
}}
function drawLine(points, color, width) {{
  const pts = points.map(p => ({{...p, t: parseDate(p.date).getTime(), value:p.sgd}}));
  const path = pts.map((p,i) => `${{i?'L':'M'}} ${{x(p.t).toFixed(2)}} ${{y(p.value).toFixed(2)}}`).join(' ');
  el('path', {{d:path, fill:'none', stroke:color, 'stroke-width':width, 'stroke-linejoin':'round', 'stroke-linecap':'round'}});
}}
drawLine(data.total, colors[0], 3.2);
let ci = 1;
for (const [name, pts] of Object.entries(data.series)) {{
  drawLine(pts, colors[ci % colors.length], 2);
  const latest = pts[pts.length - 1];
  const div = document.createElement('div');
  div.className = 'metric';
  div.innerHTML = `<strong>${{name}}</strong><span>${{latest.date}}: ${{fmt(latest.sgd)}}</span>`;
  legend.appendChild(div);
  ci++;
}}
const totalDiv = document.createElement('div');
totalDiv.className = 'metric';
totalDiv.innerHTML = `<strong>Total</strong><span>${{data.total[data.total.length-1].date}}: ${{fmt(data.total[data.total.length-1].sgd)}}</span>`;
legend.prepend(totalDiv);
el('text', {{x:L, y:22, 'font-size':16, fill:'#1d2733', 'font-weight':700}}, 'Inferred Vanguard Value Over Time, SGD');
</script>
"""
    write_html_report(REPORTS_DIR / "vanguard_inferred_value.html", "Vanguard Inferred Value", body)
    return {"points": len(total), "series": len(series), "report": str(REPORTS_DIR / "vanguard_inferred_value.html")}


if __name__ == "__main__":
    print(run())
