from __future__ import annotations

import re
from decimal import Decimal
from pathlib import Path
from typing import Any

from common import (
    BALANCE_COLUMNS,
    CONTROL_TOTAL_COLUMNS,
    ISSUE_COLUMNS,
    PROCESSED_DIR,
    TRANSACTION_COLUMNS,
    clean_description,
    converted_with_fx,
    direction_for,
    ensure_dirs,
    make_issue,
    owner_for_path,
    parse_date,
    parse_decimal,
    parser_source_file,
    source_folders_for,
    stable_id,
    write_csv,
)


PARSER_NAME = "ingest_vanguard_v1"
VANGUARD_CURRENCY = "GBP"


def vanguard_files() -> list[Path]:
    candidates: list[Path] = []
    for root in source_folders_for("vanguard"):
        candidates.extend(list(root.glob("*.xlsx")) + list(root.glob("*.Xlsx")) + list(root.glob("*.XLSX")))
    return sorted(path for path in candidates if not path.name.startswith("~$") and path.stat().st_size > 1024)


def account_from_sheet(title: str, first_cell: Any) -> tuple[str, str, str]:
    account_name = clean_description(first_cell) or clean_description(title)
    match = re.search(r"\(([^)]+)\)", title)
    if match:
        account_id = match.group(1)
    elif "Pension" in account_name:
        account_id = "VANGUARD_PERSONAL_PENSION"
    else:
        account_id = stable_id("acct", "vanguard", title)
    if "ISA" in account_name:
        account_type = "isa"
    elif "Pension" in account_name:
        account_type = "pension"
    else:
        account_type = "investment"
    return account_id, account_name, account_type


def is_transfer_candidate(details: str) -> bool:
    text = details.lower()
    return any(clue in text for clue in ["deposit", "transfer", "payment", "withdrawal", "contribution"])


def ingest_file(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    owner = owner_for_path(path)
    try:
        from openpyxl import load_workbook
    except ImportError:
        issue = make_issue(
            "parse",
            "error",
            owner,
            "vanguard",
            "ALL",
            "openpyxl is not installed, so the Vanguard workbook could not be parsed.",
            "Run with the bundled Codex Python runtime or install requirements.txt.",
            source_file=parser_source_file(path),
            key="vanguard-openpyxl-missing",
        )
        return [], [], [issue], []

    wb = load_workbook(path, read_only=True, data_only=True)
    transactions: list[dict[str, Any]] = []
    balances: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    controls: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        if ws.title == "Summary":
            continue
        first_cell = ws.cell(row=1, column=1).value
        account_id, account_name, account_type = account_from_sheet(ws.title, first_cell)
        if "NEEDS_REVIEW" in account_id:
            issues.append(
                make_issue(
                    "parse",
                    "warning",
                    "samuel",
                    "vanguard",
                    account_id,
                    "Vanguard pension account ID could not be confirmed from the truncated sheet title.",
                    "Confirm the pension account ID and update account aliases/config.",
                    source_file=parser_source_file(path),
                    key=f"vanguard-account-id-{ws.title}",
                )
            )

        in_cash_table = False
        investment_table_detected = False
        row_count = 0
        failed_rows = 0
        min_date = None
        max_date = None
        credit_sum = Decimal("0")
        debit_sum = Decimal("0")
        opening_balance = None
        closing_balance = None

        for idx, values in enumerate(ws.iter_rows(values_only=True), start=1):
            values = list(values)
            first = values[0] if len(values) > 0 else None
            second = values[1] if len(values) > 1 else None
            if first == "Date" and second == "Details":
                in_cash_table = True
                continue
            if first == "Date" and second == "InvestmentName":
                in_cash_table = False
                investment_table_detected = True
                continue
            if not in_cash_table:
                continue

            tx_date = parse_date(first)
            if tx_date is None:
                continue
            details = clean_description(second)
            amount = parse_decimal(values[2] if len(values) > 2 else None)
            balance = parse_decimal(values[3] if len(values) > 3 else None)
            if amount is None:
                failed_rows += 1
                issues.append(
                    make_issue(
                        "parse",
                        "warning",
                        owner,
                        "vanguard",
                        account_id,
                        "Vanguard cash transaction row has a date but no parseable amount.",
                        "Inspect the workbook row and update parser handling if needed.",
                        source_file=parser_source_file(path),
                        source_page=ws.title,
                        key=f"vanguard-parse-{ws.title}-{idx}",
                    )
                )
                continue

            row_count += 1
            min_date = tx_date if min_date is None or tx_date < min_date else min_date
            max_date = tx_date if max_date is None or tx_date > max_date else max_date
            if amount > 0:
                credit_sum += amount
            elif amount < 0:
                debit_sum += abs(amount)
            amount_sgd, amount_fx = converted_with_fx(amount, VANGUARD_CURRENCY, tx_date)

            transactions.append(
                {
                    "transaction_id": stable_id("tx", "vanguard", account_id, idx, tx_date, amount, details),
                    "owner": owner,
                    "institution": "vanguard",
                    "account_id": account_id,
                    "account_name": account_name,
                    "account_type": account_type,
                    "date": tx_date,
                    "posted_date": tx_date,
                    "description_raw": details,
                    "description_clean": details.lower(),
                    "merchant": "Vanguard",
                    "amount": amount,
                    "currency": VANGUARD_CURRENCY,
                    "amount_sgd": amount_sgd,
                    **amount_fx,
                    "direction": direction_for(amount),
                    "category": "investment",
                    "subcategory": "cash_transaction",
                    "is_transfer_candidate": is_transfer_candidate(details),
                    "matched_transfer_id": "",
                    "confidence_status": "confirmed",
                    "source_file": parser_source_file(path),
                    "source_page": ws.title,
                    "source_row": idx,
                    "parser_name": PARSER_NAME,
                    "parse_confidence": 0.92,
                }
            )

            if balance is not None:
                balance_sgd, balance_fx = converted_with_fx(balance, VANGUARD_CURRENCY, tx_date)
                if opening_balance is None:
                    opening_balance = balance - amount
                closing_balance = balance
                balances.append(
                    {
                        "balance_id": stable_id("bal", "vanguard", account_id, idx, tx_date, balance),
                        "owner": owner,
                        "institution": "vanguard",
                        "account_id": account_id,
                        "account_name": account_name,
                        "account_type": account_type,
                        "date": tx_date,
                        "balance": balance,
                        "currency": VANGUARD_CURRENCY,
                        "balance_sgd": balance_sgd,
                        **balance_fx,
                        "balance_type": "workbook_running_balance",
                        "confidence_status": "needs_review",
                        "source_file": parser_source_file(path),
                        "source_page": ws.title,
                        "source_row": idx,
                        "parser_name": PARSER_NAME,
                        "parse_confidence": 0.88,
                    }
                )

        issues.append(
            make_issue(
                "valuation",
                "warning",
                owner,
                "vanguard",
                account_id,
                "Vanguard cash Balance column is parsed but excluded from net worth; inferred investment values are used instead.",
                "Keep using the inferred holdings/value parser unless Vanguard provides explicit statement valuations.",
                source_file=parser_source_file(path),
                source_page=ws.title,
                status="confirmed",
                key=f"vanguard-balance-meaning-{ws.title}",
            )
        )
        if investment_table_detected:
            issues.append(
                make_issue(
                    "parse",
                "info",
                owner,
                    "vanguard",
                    account_id,
                    "Vanguard investment transaction table is normalized by ingest_vanguard_inferred.py for net worth.",
                    "No action needed unless workbook semantics change.",
                    source_file=parser_source_file(path),
                    source_page=ws.title,
                    status="confirmed",
                    key=f"vanguard-investment-table-{ws.title}",
                )
            )

        controls.append(
            {
                "control_id": stable_id("ctl", PARSER_NAME, parser_source_file(path), ws.title),
                "parser_name": PARSER_NAME,
                "source_file": parser_source_file(path),
                "owner": owner,
                "institution": "vanguard",
                "account_id": account_id,
                "file_count": 1,
                "row_count": row_count,
                "date_min": min_date,
                "date_max": max_date,
                "sum_credits": credit_sum,
                "sum_debits": debit_sum,
                "opening_balance": opening_balance,
                "closing_balance": closing_balance,
                "warning_count": len([issue for issue in issues if issue.get("account_id") == account_id]),
                "failed_row_count": failed_rows,
            }
        )

    return transactions, balances, issues, controls


def run() -> dict[str, Any]:
    ensure_dirs()
    all_transactions: list[dict[str, Any]] = []
    all_balances: list[dict[str, Any]] = []
    all_issues: list[dict[str, Any]] = []
    all_controls: list[dict[str, Any]] = []
    files = vanguard_files()

    for path in files:
        transactions, balances, issues, controls = ingest_file(path)
        all_transactions.extend(transactions)
        all_balances.extend(balances)
        all_issues.extend(issues)
        all_controls.extend(controls)

    if not files:
        all_issues.append(
            make_issue(
                "parse",
                "error",
                "unknown",
                "vanguard",
                "ALL",
                "No Vanguard XLSX files were found.",
                "Check source folder path and Vanguard export location.",
                key="vanguard-no-files",
            )
        )

    write_csv(PROCESSED_DIR / "vanguard_transactions.csv", all_transactions, TRANSACTION_COLUMNS)
    write_csv(PROCESSED_DIR / "vanguard_balances.csv", all_balances, BALANCE_COLUMNS)
    write_csv(PROCESSED_DIR / "vanguard_parse_issues.csv", all_issues, ISSUE_COLUMNS)
    write_csv(PROCESSED_DIR / "vanguard_control_totals.csv", all_controls, CONTROL_TOTAL_COLUMNS)
    return {
        "files": len(files),
        "transactions": len(all_transactions),
        "balances": len(all_balances),
        "issues": len(all_issues),
    }


if __name__ == "__main__":
    print(run())
